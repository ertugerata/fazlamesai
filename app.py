import sqlite3
import openpyxl
from flask import Flask, render_template, request, jsonify, send_file
from database import init_db, get_db_connection
import json
import os
import logging
from dotenv import load_dotenv
from datetime import datetime, date
from io import BytesIO

# --- Configuration & Logging (12 Factor) ---
load_dotenv()

# Loglama yapılandırması (stdout)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# --- Constants ---
OFFICIAL_HOLIDAYS_2025 = [
    {'date': '2025-01-01', 'description': 'Yılbaşı'},
    {'date': '2025-03-31', 'description': 'Ramazan Bayramı'},
    {'date': '2025-04-01', 'description': 'Ramazan Bayramı'},
    {'date': '2025-04-02', 'description': 'Ramazan Bayramı'},
    {'date': '2025-04-23', 'description': 'Ulusal Egemenlik'},
    {'date': '2025-05-01', 'description': 'Emek ve Dayanışma'},
    {'date': '2025-05-19', 'description': 'Gençlik ve Spor'},
    {'date': '2025-06-27', 'description': 'Kurban Bayramı'},
    {'date': '2025-06-28', 'description': 'Kurban Bayramı'},
    {'date': '2025-06-29', 'description': 'Kurban Bayramı'},
    {'date': '2025-06-30', 'description': 'Kurban Bayramı'},
    {'date': '2025-08-30', 'description': 'Zafer Bayramı'},
    {'date': '2025-09-05', 'description': 'Kurban Bayramı'},
    {'date': '2025-09-06', 'description': 'Kurban Bayramı'},
    {'date': '2025-09-07', 'description': 'Kurban Bayramı'},
    {'date': '2025-09-08', 'description': 'Kurban Bayramı'},
    {'date': '2025-10-29', 'description': 'Cumhuriyet Bayramı'},
]

app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'dev-secret-key')

# Veritabanı dosyası konfigürasyonu
DB_FILE = os.getenv('DATABASE_FILE', 'overtime.db')

# Uygulamayı başlatmadan önce veritabanının var olduğundan emin ol
if not os.path.exists(DB_FILE):
    logger.info(f"Veritabanı ({DB_FILE}) bulunamadı, oluşturuluyor...")
    init_db()

# --- Helper Functions ---
def get_days_in_month(year_month):
    year, month = map(int, year_month.split('-'))
    return (date(year, month + 1, 1) - date(year, month, 1)).days if month < 12 else 31

def get_working_days_in_month(year_month, custom_holidays):
    year, month = map(int, year_month.split('-'))
    days_in_month = get_days_in_month(year_month)
    working_days = 0
    official_holiday_dates = [h['date'] for h in OFFICIAL_HOLIDAYS_2025]

    for day in range(1, days_in_month + 1):
        current_date = date(year, month, day)
        date_str = current_date.isoformat()
        if current_date.weekday() < 5 and date_str not in custom_holidays and date_str not in official_holiday_dates:
            working_days += 1
    return working_days

def calculate_payment_for_employee(emp, year_month, logs, custom_holidays, settings):
    year, month = map(int, year_month.split('-'))
    days_in_month = get_days_in_month(year_month)

    # 1. Fetch Salary Type Config
    salary_type_id = emp['salary_type_id']
    salary_type_name = emp['salary_type_name']

    # Defaults (if no salary type linked, though we seeded defaults)
    config = {
        'include_min_wage': 0, 'include_fixed_salary': 0,
        'include_fixed_overtime_pay': 0, 'include_fixed_hours_quota': 0,
        'include_overtime_calc': 0, 'include_on_call': 0
    }

    # If we have the config columns joined in 'emp', use them
    if 'include_min_wage' in emp:
         config = {
            'include_min_wage': emp['include_min_wage'],
            'include_fixed_salary': emp['include_fixed_salary'],
            'include_fixed_overtime_pay': emp['include_fixed_overtime_pay'],
            'include_fixed_hours_quota': emp['include_fixed_hours_quota'],
            'include_overtime_calc': emp['include_overtime_calc'],
            'include_on_call': emp['include_on_call']
        }

    # 2. Categorize Logs
    weekday_day, weekday_evening, weekend_day, weekend_evening = 0, 0, 0, 0
    for day in range(1, days_in_month + 1):
        d = date(year, month, day)
        log = logs.get(d.isoformat(), {'day_hours': 0, 'evening_hours': 0})
        is_weekend = d.weekday() >= 5
        is_holiday = d.isoformat() in custom_holidays or d.isoformat() in [h['date'] for h in OFFICIAL_HOLIDAYS_2025]

        if is_weekend or is_holiday:
            weekend_day += log['day_hours']
            weekend_evening += log['evening_hours']
        else:
            weekday_day += log['day_hours']
            weekday_evening += log['evening_hours']

    # 3. Employee Constants
    fixed_salary = float(emp['fixed_salary'] or 0)
    fixed_overtime_pay = float(emp['fixed_overtime_pay'] or 0)
    try: fixed_day_hours = float(emp['fixed_day_hours'] or 0)
    except: fixed_day_hours = 0.0
    try: fixed_evening_hours = float(emp['fixed_evening_hours'] or 0)
    except: fixed_evening_hours = 0.0

    # 4. Global Settings
    day_rate = settings.get('dayRate', 0)
    evening_rate = settings.get('eveningRate', 0)
    minimum_wage = settings.get('minimumWage', 0)

    # 5. Calculation Logic
    total_payment = 0
    overtime_hours = 0
    overtime_payment = 0
    details = []

    # A. Base Pay (Min Wage + Fixed Salary)
    if config['include_min_wage']:
        total_payment += minimum_wage
        details.append(f"Asgari Ücret ({minimum_wage} TL)")

    if config['include_fixed_salary']:
        total_payment += fixed_salary
        details.append(f"Sabit Maaş ({fixed_salary} TL)")

    # B. Fixed Adds (Fixed OT Pay + Quota)
    if config['include_fixed_overtime_pay']:
        total_payment += fixed_overtime_pay
        overtime_payment += fixed_overtime_pay
        details.append(f"Sabit FM Ücreti ({fixed_overtime_pay} TL)")

    if config['include_fixed_hours_quota']:
        quota_pay = (fixed_day_hours * day_rate) + (fixed_evening_hours * evening_rate)
        total_payment += quota_pay
        overtime_payment += quota_pay
        overtime_hours += (fixed_day_hours + fixed_evening_hours)
        details.append(f"Sabit Ders: {fixed_day_hours}G + {fixed_evening_hours}A")

    # C. Variable Adds (On Call + Overtime Calc)

    # C1. On Call (Nöbet) -> Adds Weekend/Holiday work
    if config['include_on_call']:
        # Nöbet genellikle akşam tarifesinden ödenir
        on_call_pay = (weekend_day + weekend_evening) * evening_rate
        total_payment += on_call_pay
        overtime_payment += on_call_pay
        overtime_hours += (weekend_day + weekend_evening)
        details.append(f"Nöbet (H.Sonu): {weekend_day + weekend_evening}s")

    # C2. Overtime Calc (Fazla Mesai) -> Adds Weekday work (usually)
    if config['include_overtime_calc']:
        # If Min Wage is included, we deduct expected hours. Else we pay all.
        billable_weekday_day = weekday_day
        billable_weekday_evening = weekday_evening
        deducted_hours = 0

        if config['include_min_wage']:
            working_days = get_working_days_in_month(year_month, custom_holidays)
            expected_hours = working_days * 4

            # Deduct from day hours first
            deducted = min(billable_weekday_day, expected_hours)
            billable_weekday_day -= deducted
            remaining_deduction = expected_hours - deducted

            # Note: Usually we don't deduct evening hours, but if day isn't enough?
            # Standard practice in this app seems to be deducting from total or day.
            # Code `max(0, weekday_day - expected_hours)` implies only day hours are deducted.

            deducted_hours = expected_hours
            details.append(f"Düşülen Saat: {expected_hours}")

        calc_pay = (billable_weekday_day * day_rate) + (billable_weekday_evening * evening_rate)

        # If 'On Call' is NOT active, maybe weekends count as FM here?
        # Standard 'Asgari + FM' (MinWage=1, Calc=1, OnCall=0) included weekends in existing code.
        if not config['include_on_call']:
             calc_pay += (weekend_day * evening_rate) + (weekend_evening * evening_rate)
             overtime_hours += (weekend_day + weekend_evening)

        total_payment += calc_pay
        overtime_payment += calc_pay
        overtime_hours += (billable_weekday_day + billable_weekday_evening) # Only add billed hours? Or all worked extra hours?
        # For display, let's show billable.

        if calc_pay > 0:
            details.append("Hesaplanan FM Eklendi")

    # Helper for branch
    try: branch = emp['branch']
    except: branch = ''

    return {
        'emp_id': emp['id'], 'name': emp['name'], 'empId': emp['emp_id'], 'branch': branch,
        'paymentType': salary_type_name, # Display Name
        'fixedSalary': fixed_salary,
        'fixedOvertimePay': fixed_overtime_pay,
        'fixedDayHours': fixed_day_hours, 'fixedEveningHours': fixed_evening_hours,
        'weekdayDayHours': weekday_day, 'weekdayEveningHours': weekday_evening,
        'weekendDayHours': weekend_day, 'weekendEveningHours': weekend_evening,
        'totalHours': weekday_day + weekday_evening + weekend_day + weekend_evening,
        'overtimeHours': overtime_hours,
        'overtimePayment': overtime_payment,
        'totalPayment': total_payment,
        'minimumWage': minimum_wage if config['include_min_wage'] else 0,
        'calculationDetails': " + ".join(details)
    }

# --- Ana Sayfa ---
@app.route('/')
def index():
    return render_template('index.html')

# --- Ayarlar API ---
@app.route('/api/settings', methods=['GET'])
def get_settings():
    conn = get_db_connection()
    settings = conn.execute('SELECT key, value FROM settings').fetchall()
    conn.close()
    return jsonify({s['key']: s['value'] for s in settings})

@app.route('/api/settings', methods=['POST'])
def update_settings():
    data = request.json
    conn = get_db_connection()
    for key, value in data.items():
        conn.execute("UPDATE settings SET value = ? WHERE key = ?", (str(value), key))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Ayarlar güncellendi.'})

# --- Maaş Opsiyonları (Salary Types) API ---
@app.route('/api/salary_types', methods=['GET'])
def get_salary_types():
    conn = get_db_connection()
    types = conn.execute('SELECT * FROM salary_types').fetchall()
    conn.close()
    return jsonify([dict(t) for t in types])

@app.route('/api/salary_types', methods=['POST'])
def add_salary_type():
    data = request.json
    name = data.get('name')
    if not name: return jsonify({'error': 'İsim zorunludur.'}), 400

    conn = get_db_connection()
    cursor = conn.execute("""
        INSERT INTO salary_types (name, include_min_wage, include_fixed_salary, include_fixed_overtime_pay, include_fixed_hours_quota, include_overtime_calc, include_on_call)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (name, data.get('include_min_wage',0), data.get('include_fixed_salary',0),
          data.get('include_fixed_overtime_pay',0), data.get('include_fixed_hours_quota',0),
          data.get('include_overtime_calc',0), data.get('include_on_call',0)))
    conn.commit()
    new_id = cursor.lastrowid
    conn.close()
    return jsonify({'id': new_id, 'message': 'Maaş opsiyonu eklendi.'}), 201

@app.route('/api/salary_types/<int:id>', methods=['DELETE'])
def delete_salary_type(id):
    conn = get_db_connection()
    # Check usage
    count = conn.execute("SELECT COUNT(*) FROM employees WHERE salary_type_id = ?", (id,)).fetchone()[0]
    if count > 0:
        conn.close()
        return jsonify({'error': 'Bu opsiyon çalışanlar tarafından kullanılıyor, silinemez.'}), 400

    conn.execute("DELETE FROM salary_types WHERE id = ?", (id,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Silindi.'})

# --- Çalışanlar API ---
@app.route('/api/employees', methods=['GET'])
def get_employees():
    conn = get_db_connection()
    # Join with salary_types to get name and config
    query = """
        SELECT e.*, s.name as salary_type_name,
               s.include_min_wage, s.include_fixed_salary, s.include_fixed_overtime_pay,
               s.include_fixed_hours_quota, s.include_overtime_calc, s.include_on_call
        FROM employees e
        LEFT JOIN salary_types s ON e.salary_type_id = s.id
        ORDER BY e.name
    """
    employees = conn.execute(query).fetchall()
    conn.close()
    return jsonify([dict(emp) for emp in employees])

@app.route('/api/employees', methods=['POST'])
def add_employee():
    data = request.json
    name, emp_id, branch = data.get('name'), data.get('emp_id'), data.get('branch')
    if not name: return jsonify({'error': 'İsim zorunludur.'}), 400

    conn = get_db_connection()
    # Get default salary type (first one?)
    default_type = conn.execute("SELECT id FROM salary_types LIMIT 1").fetchone()
    salary_type_id = default_type['id'] if default_type else None

    cursor = conn.execute("INSERT INTO employees (name, emp_id, branch, salary_type_id) VALUES (?, ?, ?, ?)", (name, emp_id, branch, salary_type_id))
    conn.commit()
    new_id = cursor.lastrowid

    # Fetch full object to return
    emp = conn.execute("""
        SELECT e.*, s.name as salary_type_name
        FROM employees e LEFT JOIN salary_types s ON e.salary_type_id = s.id
        WHERE e.id = ?
    """, (new_id,)).fetchone()
    conn.close()

    return jsonify(dict(emp)), 201

@app.route('/api/employees/<int:id>', methods=['PUT'])
def update_employee(id):
    data = request.json

    # Fields that can be updated
    salary_type_id = data.get('salary_type_id')
    fixed_salary = data.get('fixed_salary')
    fixed_overtime_pay = data.get('fixed_overtime_pay')
    fixed_day_hours = data.get('fixed_day_hours')
    fixed_evening_hours = data.get('fixed_evening_hours')
    branch = data.get('branch')

    conn = get_db_connection()

    conn.execute("""
        UPDATE employees SET
        salary_type_id = ?, fixed_salary = ?, fixed_overtime_pay = ?,
        fixed_day_hours = ?, fixed_evening_hours = ?, branch = ?
        WHERE id = ?
    """, (salary_type_id, fixed_salary, fixed_overtime_pay, fixed_day_hours, fixed_evening_hours, branch, id))

    conn.commit()
    conn.close()
    return jsonify({'message': 'Çalışan güncellendi.'})

@app.route('/api/employees/bulk', methods=['POST'])
def add_bulk_employees():
    employees_data = request.json.get('employees', [])
    if not employees_data: return jsonify({'error': 'Çalışan listesi boş.'}), 400

    conn = get_db_connection()
    # Get default salary type
    default_type = conn.execute("SELECT id FROM salary_types LIMIT 1").fetchone()
    salary_type_id = default_type['id'] if default_type else None

    conn.executemany("INSERT INTO employees (name, emp_id, branch, salary_type_id) VALUES (?, ?, ?, ?)",
                     [(e.get('name'), e.get('emp_id'), e.get('branch'), salary_type_id) for e in employees_data])
    conn.commit()
    conn.close()
    return jsonify({'message': f'{len(employees_data)} çalışan eklendi.'}), 201

@app.route('/api/employees/template', methods=['GET'])
def download_employee_template():
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Çalışan Ekleme Şablonu"

    headers = [
        'Ad Soyad', 'Çalışan No', 'Branş', 'Ödeme Tipi (Opsiyon Adı)',
        'Sabit Maaş', 'Sabit FM Ücreti',
        'Gündüz Sabit Ders', 'Akşam Sabit Ders'
    ]
    sheet.append(headers)

    # Add validation info as a second sheet
    conn = get_db_connection()
    types = conn.execute("SELECT name FROM salary_types").fetchall()
    conn.close()

    info_sheet = wb.create_sheet("Bilgi")
    info_sheet.append(["Geçerli Ödeme Tipleri (Kopyalayıp Yapıştırın)"])
    for t in types:
        info_sheet.append([t['name']])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='calisan_ekleme_sablonu.xlsx')

@app.route('/api/employees/upload', methods=['POST'])
def upload_employees():
    file = request.files.get('file')
    if not file: return jsonify({'error': 'Dosya bulunamadı.'}), 400

    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    employees_to_add = []

    conn = get_db_connection()
    # Load all salary types into a map: Name -> ID
    salary_types = {row['name']: row['id'] for row in conn.execute("SELECT name, id FROM salary_types").fetchall()}
    default_type_id = list(salary_types.values())[0] if salary_types else None
    conn.close()

    def safe_float(val):
        try: return float(val)
        except (ValueError, TypeError): return 0.0

    # Headers in row 1
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue

        name = row[0]
        emp_id = str(row[1]) if row[1] else None
        branch = row[2] if len(row) > 2 else None

        # Salary Type Lookup
        payment_type_str = row[3] if len(row) > 3 else None
        salary_type_id = salary_types.get(payment_type_str, default_type_id)

        fixed_salary = safe_float(row[4] if len(row) > 4 else 0)
        fixed_overtime_pay = safe_float(row[5] if len(row) > 5 else 0)
        fixed_day_hours = safe_float(row[6] if len(row) > 6 else 0)
        fixed_evening_hours = safe_float(row[7] if len(row) > 7 else 0)

        employees_to_add.append({
            'name': name, 'emp_id': emp_id, 'branch': branch,
            'salary_type_id': salary_type_id,
            'fixed_salary': fixed_salary,
            'fixed_overtime_pay': fixed_overtime_pay,
            'fixed_day_hours': fixed_day_hours,
            'fixed_evening_hours': fixed_evening_hours
        })

    if not employees_to_add: return jsonify({'message': 'Eklenecek çalışan bulunamadı.'})

    conn = get_db_connection()
    conn.executemany("""
        INSERT INTO employees (
            name, emp_id, branch, salary_type_id,
            fixed_salary, fixed_overtime_pay,
            fixed_day_hours, fixed_evening_hours
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, [(
        e['name'], e['emp_id'], e['branch'], e['salary_type_id'],
        e['fixed_salary'], e['fixed_overtime_pay'],
        e['fixed_day_hours'], e['fixed_evening_hours']
    ) for e in employees_to_add])
    conn.commit()
    conn.close()
    return jsonify({'message': f'{len(employees_to_add)} çalışan Excel\'den eklendi.'})

@app.route('/api/employees/<int:id>', methods=['DELETE'])
def delete_employee(id):
    conn = get_db_connection()
    conn.execute('DELETE FROM employees WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Çalışan silindi.'})

# --- Tatiller API ---
@app.route('/api/holidays', methods=['GET'])
def get_holidays():
    conn = get_db_connection()
    holidays = conn.execute('SELECT date FROM holidays ORDER BY date').fetchall()
    conn.close()
    return jsonify([h['date'] for h in holidays])

@app.route('/api/holidays', methods=['POST'])
def add_holiday():
    date = request.json.get('date')
    if not date: return jsonify({'error': 'Tarih zorunludur.'}), 400
    conn = get_db_connection()
    try:
        conn.execute("INSERT INTO holidays (date) VALUES (?)", (date,))
        conn.commit()
    except sqlite3.IntegrityError:
        return jsonify({'error': 'Bu tarih zaten ekli.'}), 409
    finally: conn.close()
    return jsonify({'message': 'Tatil eklendi.'}), 201

@app.route('/api/holidays/<string:date>', methods=['DELETE'])
def delete_holiday(date):
    conn = get_db_connection()
    conn.execute('DELETE FROM holidays WHERE date = ?', (date,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Tatil silindi.'})

# --- Çalışma Saatleri API ---
@app.route('/api/worklogs/<string:year_month>', methods=['GET'])
def get_work_logs(year_month):
    conn = get_db_connection()
    logs_data = conn.execute(
        "SELECT employee_id, date, day_hours, evening_hours, sunday_reason FROM work_logs WHERE strftime('%Y-%m', date) = ?",
        (year_month,)
    ).fetchall()
    conn.close()
    result = {}
    for log in logs_data:
        emp_id = str(log['employee_id'])
        if emp_id not in result: result[emp_id] = {}
        result[emp_id][log['date']] = {'day': log['day_hours'], 'evening': log['evening_hours'], 'reason': log['sunday_reason']}
    return jsonify(result)

@app.route('/api/worklogs', methods=['POST'])
def update_work_log():
    data = request.json
    emp_id, date, log_type, value, reason = data.get('empId'), data.get('date'), data.get('type'), data.get('value', 0), data.get('reason')
    if not all([emp_id, date, log_type]): return jsonify({'error': 'Eksik parametre.'}), 400

    try:
        value = int(value or 0)
    except (ValueError, TypeError):
        return jsonify({'error': 'Saat değeri tam sayı olmalıdır.'}), 400

    field = 'day_hours' if log_type == 'day' else 'evening_hours'
    conn = get_db_connection()
    existing_log = conn.execute("SELECT id FROM work_logs WHERE employee_id = ? AND date = ?", (emp_id, date)).fetchone()
    if existing_log:
        query = f"UPDATE work_logs SET {field} = ?"
        params = [value]
        if reason is not None:
            query += ", sunday_reason = ?"
            params.append(reason)
        query += " WHERE id = ?"
        params.append(existing_log['id'])
        conn.execute(query, tuple(params))
    else:
        other_field = 'evening_hours' if log_type == 'day' else 'day_hours'
        conn.execute(f"INSERT INTO work_logs (employee_id, date, {field}, {other_field}, sunday_reason) VALUES (?, ?, ?, 0, ?)",
                     (emp_id, date, value, reason))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Çalışma saati güncellendi.'})

@app.route('/api/worklogs/upload', methods=['POST'])
def upload_worklogs():
    file = request.files.get('file')
    if not file: return jsonify({'error': 'Dosya bulunamadı.'}), 400

    conn = get_db_connection()
    employees = {e['name']: e['id'] for e in conn.execute("SELECT id, name FROM employees").fetchall()}
    wb = openpyxl.load_workbook(file)

    def process_sheet(sheet_name, type_):
        if sheet_name not in wb.sheetnames: return
        sheet = wb[sheet_name]
        headers = [cell.value for cell in sheet[1]]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            emp_name = row[0]
            if emp_name in employees:
                emp_id = employees[emp_name]
                for i, hours in enumerate(row[1:], 1):
                    if hours:
                        try:
                            hours_val = int(float(hours))
                            if hours_val > 0:
                                date_str = headers[i]
                                update_work_log_db(conn, emp_id, date_str, type_, hours_val)
                        except (ValueError, TypeError):
                            continue

    process_sheet('Gündüz Mesaisi', 'day')
    process_sheet('Akşam Mesaisi', 'evening')

    conn.commit()
    conn.close()
    return jsonify({'message': 'Çalışma saatleri yüklendi.'})

def update_work_log_db(conn, emp_id, date_str, log_type, value):
    field = 'day_hours' if log_type == 'day' else 'evening_hours'
    existing = conn.execute("SELECT id FROM work_logs WHERE employee_id = ? AND date = ?", (emp_id, date_str)).fetchone()
    if existing:
        conn.execute(f"UPDATE work_logs SET {field} = ? WHERE id = ?", (value, existing['id']))
    else:
        other_field = 'evening_hours' if log_type == 'day' else 'day_hours'
        conn.execute(f"INSERT INTO work_logs (employee_id, date, {field}, {other_field}) VALUES (?, ?, ?, 0)",
                     (emp_id, date_str, value))

@app.route('/api/worklogs/template/<string:year_month>', methods=['GET'])
def download_worklog_template(year_month):
    conn = get_db_connection()
    employees = conn.execute("SELECT name FROM employees ORDER BY name").fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    day_sheet = wb.active
    day_sheet.title = "Gündüz Mesaisi"
    evening_sheet = wb.create_sheet("Akşam Mesaisi")

    headers = ['Ad Soyad'] + [f"{year_month}-{str(day).zfill(2)}" for day in range(1, get_days_in_month(year_month) + 1)]
    day_sheet.append(headers)
    evening_sheet.append(headers)

    for emp in employees:
        day_sheet.append([emp['name']])
        evening_sheet.append([emp['name']])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f'calisma-sablonu-{year_month}.xlsx')


# --- Raporlama API ---
@app.route('/api/report/<string:year_month>', methods=['GET'])
def get_report(year_month):
    conn = get_db_connection()
    # Fetch employees with their salary type config
    employees = conn.execute("""
        SELECT e.*, s.name as salary_type_name,
               s.include_min_wage, s.include_fixed_salary, s.include_fixed_overtime_pay,
               s.include_fixed_hours_quota, s.include_overtime_calc, s.include_on_call
        FROM employees e
        LEFT JOIN salary_types s ON e.salary_type_id = s.id
    """).fetchall()

    logs_data = conn.execute("SELECT * FROM work_logs WHERE strftime('%Y-%m', date) = ?", (year_month,)).fetchall()
    holidays = [h['date'] for h in conn.execute("SELECT date FROM holidays").fetchall()]
    settings = {s['key']: float(s['value']) for s in conn.execute("SELECT key, value FROM settings").fetchall()}
    conn.close()

    logs_by_emp = {}
    for log in logs_data:
        emp_id = log['employee_id']
        if emp_id not in logs_by_emp: logs_by_emp[emp_id] = {}
        logs_by_emp[emp_id][log['date']] = log

    report = [calculate_payment_for_employee(dict(emp), year_month, logs_by_emp.get(emp['id'], {}), holidays, settings) for emp in employees]
    return jsonify(report)

@app.route('/api/report/export/<string:year_month>', methods=['GET'])
def export_report(year_month):
    report_data = get_report(year_month).get_json()
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Maaş Raporu"

    headers = ['Ad Soyad', 'Çalışan No', 'Branş', 'Ödeme Tipi', 'Sabit Maaş', 'Asgari Ücret',
               'Fazla Mesai Saati', 'Fazla Mesai Ödemesi', 'Toplam Hakediş', 'Açıklama']
    sheet.append(headers)

    for data in report_data:
        sheet.append([
            data['name'], data['empId'], data.get('branch', ''), data['paymentType'],
            f"{data['fixedSalary']:.2f}", f"{data['minimumWage']:.2f}",
            data['overtimeHours'], f"{data['overtimePayment']:.2f}",
            f"{data['totalPayment']:.2f}", data['calculationDetails']
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f'maas-raporu-{year_month}.xlsx')

@app.route('/api/export_all', methods=['GET'])
def export_all_data():
    conn = get_db_connection()

    employees = conn.execute("""
        SELECT e.id, e.name, e.emp_id, e.branch, s.name as salary_type,
               e.fixed_salary, e.fixed_overtime_pay, e.fixed_day_hours, e.fixed_evening_hours
        FROM employees e LEFT JOIN salary_types s ON e.salary_type_id = s.id
        ORDER BY e.name
    """).fetchall()

    work_logs = conn.execute("""
        SELECT e.name, w.date, w.day_hours, w.evening_hours, w.sunday_reason
        FROM work_logs w JOIN employees e ON w.employee_id = e.id
        ORDER BY e.name, w.date
    """).fetchall()
    holidays = conn.execute("SELECT date FROM holidays ORDER BY date").fetchall()
    settings = conn.execute("SELECT key, value FROM settings").fetchall()

    conn.close()

    wb = openpyxl.Workbook()

    # Çalışanlar Sayfası
    ws_employees = wb.active
    ws_employees.title = "Çalışanlar"
    ws_employees.append(['ID', 'Ad Soyad', 'Çalışan No', 'Branş', 'Ödeme Tipi', 'Sabit Maaş', 'Sabit FM Ücreti', 'Sabit Gündüz Ders', 'Sabit Akşam Ders'])
    for emp in employees:
        ws_employees.append([emp['id'], emp['name'], emp['emp_id'], emp['branch'], emp['salary_type'], emp['fixed_salary'], emp['fixed_overtime_pay'], emp['fixed_day_hours'], emp['fixed_evening_hours']])

    # Diğer sayfalar aynı...
    ws_worklogs = wb.create_sheet("Çalışma Saatleri")
    ws_worklogs.append(['Ad Soyad', 'Tarih', 'Gündüz Saati', 'Akşam Saati', 'Pazar Gerekçesi'])
    for log in work_logs:
        ws_worklogs.append([log['name'], log['date'], log['day_hours'], log['evening_hours'], log['sunday_reason']])

    ws_holidays = wb.create_sheet("Tatiller")
    ws_holidays.append(['Tarih'])
    for holiday in holidays:
        ws_holidays.append([holiday['date']])

    ws_settings = wb.create_sheet("Ayarlar")
    ws_settings.append(['Ayar', 'Değer'])
    for setting in settings:
        ws_settings.append([setting['key'], setting['value']])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(output, as_attachment=True, download_name=f'fazla_mesai_yedek_{timestamp}.xlsx')

if __name__ == '__main__':
    port = int(os.getenv('PORT', 5000))
    debug_mode = os.getenv('FLASK_DEBUG', 'True').lower() in ('true', '1', 't')
    app.run(debug=debug_mode, port=port, host='0.0.0.0')
